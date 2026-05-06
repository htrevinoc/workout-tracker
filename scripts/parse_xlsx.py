"""
Parsea data/Hypertrophy_Tracker_v2.xlsx -> data/seed.json

Estructura del XLSX (resumen):
- PROGRAM: metadatos del plan (fecha inicio, peso inicial/objetivo)
- UPPER PUSH / LOWER / UPPER PULL: por hoja
    fila 2: nombres de ejercicios (cols C..)
    fila 3: target (ej "3x6-8")
    fila 4: descanso (ej "90s")
    a partir de fila 5: bloques semanales de 4 filas:
        fila A: "S{N}" en col A, "Peso" en col B, peso por ejercicio en cols C..
        fila B: "Set 1" en col B, reps en cols C..
        fila C: "Set 2"
        fila D: "Set 3"
        Penultima col: RPE de la sesion
        Ultima col: Notas
- PROGRESS: peso corporal semanal y progresion de lifts clave (no se exporta a seed)

Notacion de peso encontrada:
- "14.5x14.5"  -> mancuerna 14.5 kg en cada mano  (mode: per_side)
- "8+8"        -> dos mancuernas, suma 8+8        (mode: pair_total)
- 30 (number)  -> peso total (barra/mancuerna)    (mode: barbell o per_side, dejamos barbell por default si es number)
- celda vacia  -> sin dato

El JSON resultante tiene catalogo de ejercicios + rutinas + sesiones inferidas
de las semanas con datos. Cada sesion tiene N session_sets.
"""

from __future__ import annotations

import json
import re
import uuid
from dataclasses import dataclass, asdict, field
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "data" / "Hypertrophy_Tracker_v2.xlsx"
OUT_PATH = ROOT / "data" / "seed.json"


# ---------- helpers ----------


def slugify(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def stable_id(*parts: str) -> str:
    """Deterministic UUID5 so re-running the script produces the same ids."""
    namespace = uuid.UUID("00000000-0000-0000-0000-000000000001")
    return str(uuid.uuid5(namespace, "|".join(parts)))


def parse_target_reps(target: str | None) -> tuple[int, int, int]:
    """
    "3x6-8"      -> (3, 6, 8)
    "3x8-10/leg" -> (3, 8, 10)
    "3x10-12"    -> (3, 10, 12)
    fallback     -> (3, 8, 12)
    """
    if not target:
        return (3, 8, 12)
    m = re.match(r"\s*(\d+)\s*x\s*(\d+)\s*(?:-\s*(\d+))?", str(target))
    if not m:
        return (3, 8, 12)
    sets = int(m.group(1))
    rmin = int(m.group(2))
    rmax = int(m.group(3) or m.group(2))
    return (sets, rmin, rmax)


def parse_rest_seconds(rest: str | None) -> int:
    if not rest:
        return 90
    m = re.search(r"(\d+)\s*s", str(rest))
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)\s*min", str(rest))
    if m:
        return int(m.group(1)) * 60
    try:
        return int(rest)
    except (TypeError, ValueError):
        return 90


def parse_weight_cell(cell: Any) -> dict | None:
    """
    Returns {"value": float, "mode": "per_side|pair_total|barbell"} or None.

    Examples:
      "14.5x14.5" -> per_side 14.5
      "10x10"     -> per_side 10
      "8+8"       -> pair_total 16  (suma de las dos mancuernas)
      14.5        -> per_side 14.5  (numero suelto en col de mancuerna; ambiguo, asumimos per_side)
      30          -> barbell 30     (numero suelto >= 25 sugiere barra/peso total)
      None        -> None
    """
    if cell is None or cell == "":
        return None
    if isinstance(cell, (int, float)):
        val = float(cell)
        # Heuristic: numbers >= 25 are likely barbell totals; below, probably DB per-side
        mode = "barbell" if val >= 25 else "per_side"
        return {"value": val, "mode": mode}
    s = str(cell).strip().lower().replace(" ", "")
    # "AxA" pattern (mancuerna por mano, ambos lados iguales)
    m = re.match(r"^(\d+(?:\.\d+)?)x(\d+(?:\.\d+)?)$", s)
    if m:
        a, b = float(m.group(1)), float(m.group(2))
        if a == b:
            return {"value": a, "mode": "per_side"}
        # asimetricas: guardamos el max
        return {"value": max(a, b), "mode": "per_side", "asymmetric": [a, b]}
    # "A+A" pattern (suma de mancuernas)
    m = re.match(r"^(\d+(?:\.\d+)?)\+(\d+(?:\.\d+)?)$", s)
    if m:
        a, b = float(m.group(1)), float(m.group(2))
        return {"value": a + b, "mode": "pair_total", "components": [a, b]}
    # numero como string
    try:
        val = float(s)
        mode = "barbell" if val >= 25 else "per_side"
        return {"value": val, "mode": mode}
    except ValueError:
        return None


def parse_reps_cell(cell: Any) -> int | None:
    if cell is None or cell == "":
        return None
    if isinstance(cell, (int, float)):
        return int(cell)
    s = str(cell).strip()
    m = re.match(r"^(\d+)", s)
    if m:
        return int(m.group(1))
    return None


# ---------- model ----------


@dataclass
class Exercise:
    id: str
    name_es: str
    name_en: str
    muscle_groups: list[str]
    equipment: str
    default_weight_mode: str
    default_target_reps_min: int
    default_target_reps_max: int
    default_rest_seconds: int
    is_otf: bool = False


@dataclass
class Routine:
    id: str
    name_es: str
    name_en: str
    color: str | None = None


@dataclass
class RoutineExercise:
    id: str
    routine_id: str
    exercise_id: str
    order_in_routine: int
    target_sets: int
    target_reps_min: int
    target_reps_max: int
    rest_seconds: int


@dataclass
class Session:
    id: str
    routine_id: str
    started_at: str
    week_number: int
    location: str = "home"
    notes: str | None = None
    session_rpe: float | None = None


@dataclass
class SessionSet:
    id: str
    session_id: str
    exercise_id: str
    set_number: int
    weight_value: float | None = None
    weight_unit: str = "kg"
    weight_mode: str | None = None
    reps: int | None = None
    rpe: float | None = None


@dataclass
class Seed:
    meta: dict = field(default_factory=dict)
    exercises: list[Exercise] = field(default_factory=list)
    routines: list[Routine] = field(default_factory=list)
    routine_exercises: list[RoutineExercise] = field(default_factory=list)
    sessions: list[Session] = field(default_factory=list)
    session_sets: list[SessionSet] = field(default_factory=list)
    bodyweight: list[dict] = field(default_factory=list)


# ---------- exercise metadata heuristics ----------

EQUIPMENT_KEYWORDS = [
    ("barbell", "barbell"),
    ("dumbbell", "dumbbell"),
    ("db ", "dumbbell"),
    (" db", "dumbbell"),
    ("machine", "machine"),
    ("cable", "cable"),
    ("band", "band"),
]

MUSCLE_KEYWORDS = [
    ("bench", ["chest", "triceps"]),
    ("press", ["chest", "shoulders"]),
    ("shoulder", ["shoulders"]),
    ("lateral raise", ["shoulders"]),
    ("tricep", ["triceps"]),
    ("squat", ["quads", "glutes"]),
    ("split squat", ["quads", "glutes"]),
    ("goblet", ["quads", "glutes"]),
    ("deadlift", ["hamstrings", "glutes", "back"]),
    ("calf", ["calves"]),
    ("row", ["back", "biceps"]),
    ("pullover", ["back", "chest"]),
    ("reverse fly", ["shoulders", "back"]),
    ("curl", ["biceps"]),
    ("hammer", ["biceps"]),
]


def infer_equipment(name: str) -> str:
    low = " " + name.lower() + " "
    for kw, eq in EQUIPMENT_KEYWORDS:
        if kw in low:
            return eq
    return "other"


def infer_muscles(name: str) -> list[str]:
    low = name.lower()
    found: list[str] = []
    for kw, muscles in MUSCLE_KEYWORDS:
        if kw in low:
            for m in muscles:
                if m not in found:
                    found.append(m)
    return found or ["other"]


def infer_default_mode(equipment: str) -> str:
    if equipment == "barbell":
        return "barbell"
    if equipment == "dumbbell":
        return "per_side"
    return "barbell"


# Translation table for known exercises (es <- en mostly the same; just polish)
NAME_ES_OVERRIDES = {
    "Barbell Bench Press": "Press de banca",
    "Incline Dumbbell Press": "Press inclinado con mancuernas",
    "Seated DB Shoulder Press": "Press militar sentado con mancuernas",
    "DB Lateral Raises": "Elevaciones laterales con mancuernas",
    "DB Tricep Overhead Ext": "Extension de triceps sobre la cabeza",
    "Barbell Back Squat": "Sentadilla con barra",
    "Barbell Romanian Deadlift": "Peso muerto rumano con barra",
    "Bulgarian Split Squat (DB)": "Sentadilla bulgara con mancuernas",
    "DB Goblet Squat": "Sentadilla goblet con mancuerna",
    "Single-Leg Calf Raise (DB)": "Elevacion de pantorrilla a una pierna",
    "DB Bent-Over Row": "Remo inclinado con mancuernas",
    "Single-Arm DB Row": "Remo a una mano con mancuerna",
    "DB Pullover": "Pullover con mancuerna",
    "Incline Reverse DB Fly": "Aperturas inversas inclinadas",
    "DB Hammer Curls": "Curl martillo",
    "Barbell Curl": "Curl con barra",
}


# ---------- main ----------


def main() -> None:
    if not XLSX_PATH.exists():
        raise SystemExit(f"No se encontro: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    seed = Seed()

    # --- PROGRAM sheet: metadata
    program = wb["PROGRAM"]
    start_date_value = program["B4"].value
    weight_initial = program["B5"].value
    weight_target = program["B6"].value

    if isinstance(start_date_value, str):
        start_date = datetime.fromisoformat(start_date_value).date()
    elif isinstance(start_date_value, (datetime, date)):
        start_date = start_date_value if isinstance(start_date_value, date) else start_date_value.date()
    else:
        start_date = date(2026, 3, 23)

    seed.meta = {
        "source_file": XLSX_PATH.name,
        "start_date": start_date.isoformat(),
        "weight_initial_kg": weight_initial,
        "weight_target_kg": weight_target,
        "generated_at": datetime.now(UTC).isoformat(),
    }

    # --- workout sheets
    workout_sheets = ["UPPER PUSH", "LOWER", "UPPER PULL"]
    routine_colors = {"UPPER PUSH": "#ef4444", "LOWER": "#22c55e", "UPPER PULL": "#3b82f6"}
    weekday_offset = {"UPPER PUSH": 0, "LOWER": 2, "UPPER PULL": 4}  # Mon, Wed, Fri

    exercises_by_name: dict[str, Exercise] = {}

    for sheet_name in workout_sheets:
        ws = wb[sheet_name]
        # Determine RPE / Notas columns: they are the last two non-empty in row 2
        header_row = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        # Exercise columns start at C (col 3); end before RPE/Notas
        # Find indexes (1-based) of "RPE" and "Notas"
        rpe_col = next((i + 1 for i, v in enumerate(header_row) if v == "RPE"), None)
        notes_col = next((i + 1 for i, v in enumerate(header_row) if v == "Notas"), None)
        last_exercise_col = (rpe_col or len(header_row) + 1) - 1
        exercise_cols = list(range(3, last_exercise_col + 1))

        # Routine
        routine_id = stable_id("routine", sheet_name)
        seed.routines.append(
            Routine(
                id=routine_id,
                name_es=sheet_name.title(),
                name_en=sheet_name.title(),
                color=routine_colors.get(sheet_name),
            )
        )

        # Exercises in this sheet
        targets_row = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
        rests_row = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]

        sheet_exercises: list[tuple[int, Exercise, RoutineExercise]] = []
        for order, col in enumerate(exercise_cols):
            name = header_row[col - 1]
            if not name:
                continue
            name = str(name).strip()
            if name in exercises_by_name:
                ex = exercises_by_name[name]
            else:
                equipment = infer_equipment(name)
                muscles = infer_muscles(name)
                t_sets, t_min, t_max = parse_target_reps(targets_row[col - 1])
                rest = parse_rest_seconds(rests_row[col - 1])
                ex = Exercise(
                    id=stable_id("exercise", name),
                    name_es=NAME_ES_OVERRIDES.get(name, name),
                    name_en=name,
                    muscle_groups=muscles,
                    equipment=equipment,
                    default_weight_mode=infer_default_mode(equipment),
                    default_target_reps_min=t_min,
                    default_target_reps_max=t_max,
                    default_rest_seconds=rest,
                )
                exercises_by_name[name] = ex
                seed.exercises.append(ex)

            t_sets, t_min, t_max = parse_target_reps(targets_row[col - 1])
            rest = parse_rest_seconds(rests_row[col - 1])
            re_obj = RoutineExercise(
                id=stable_id("routine_exercise", routine_id, ex.id),
                routine_id=routine_id,
                exercise_id=ex.id,
                order_in_routine=order,
                target_sets=t_sets,
                target_reps_min=t_min,
                target_reps_max=t_max,
                rest_seconds=rest,
            )
            seed.routine_exercises.append(re_obj)
            sheet_exercises.append((col, ex, re_obj))

        # Weekly blocks: starting at row 5, each block is 4 rows (Peso, Set 1, Set 2, Set 3)
        block_size = 4
        max_row = ws.max_row
        row = 5
        while row + block_size - 1 <= max_row:
            week_label = ws.cell(row=row, column=1).value
            if not week_label:
                row += block_size
                continue
            week_match = re.match(r"S(\d+)", str(week_label))
            if not week_match:
                row += block_size
                continue
            week_number = int(week_match.group(1))

            # Check if this week has any data
            has_data = False
            session_sets_for_week: list[SessionSet] = []
            week_rpe = None
            week_notes = None
            session_id = stable_id("session", routine_id, str(week_number))

            for col, ex, _re_obj in sheet_exercises:
                weight_cell = ws.cell(row=row, column=col).value
                weight_parsed = parse_weight_cell(weight_cell)

                for set_idx in range(1, 4):
                    reps_cell = ws.cell(row=row + set_idx, column=col).value
                    reps = parse_reps_cell(reps_cell)
                    if weight_parsed is None and reps is None:
                        continue
                    has_data = True
                    set_obj = SessionSet(
                        id=stable_id("session_set", session_id, ex.id, str(set_idx)),
                        session_id=session_id,
                        exercise_id=ex.id,
                        set_number=set_idx,
                        weight_value=weight_parsed["value"] if weight_parsed else None,
                        weight_unit="kg",
                        weight_mode=weight_parsed["mode"] if weight_parsed else None,
                        reps=reps,
                    )
                    session_sets_for_week.append(set_obj)

            if rpe_col:
                rpe_val = ws.cell(row=row, column=rpe_col).value
                if rpe_val is not None:
                    try:
                        week_rpe = float(rpe_val)
                        has_data = True
                    except (TypeError, ValueError):
                        pass
            if notes_col:
                notes_val = ws.cell(row=row, column=notes_col).value
                if notes_val:
                    week_notes = str(notes_val)
                    has_data = True

            if has_data:
                started_at = datetime.combine(
                    start_date + timedelta(weeks=week_number - 1, days=weekday_offset[sheet_name]),
                    datetime.min.time(),
                ).isoformat()
                seed.sessions.append(
                    Session(
                        id=session_id,
                        routine_id=routine_id,
                        started_at=started_at,
                        week_number=week_number,
                        location="home",
                        notes=week_notes,
                        session_rpe=week_rpe,
                    )
                )
                seed.session_sets.extend(session_sets_for_week)

            row += block_size

    # --- PROGRESS sheet: bodyweight log
    if "PROGRESS" in wb.sheetnames:
        ws = wb["PROGRESS"]
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if not row or row[0] is None:
                continue
            week, date_val, weight_val = row[0], row[1], row[2]
            if weight_val is None:
                continue
            if isinstance(date_val, datetime):
                d = date_val.date().isoformat()
            elif isinstance(date_val, date):
                d = date_val.isoformat()
            elif isinstance(date_val, str) and date_val:
                d = date_val
            else:
                d = (start_date + timedelta(weeks=int(week) - 1)).isoformat()
            seed.bodyweight.append({
                "id": stable_id("bodyweight", str(week), d),
                "date": d,
                "weight_value": float(weight_val),
                "weight_unit": "kg",
            })

    # --- write JSON
    out = {
        "meta": seed.meta,
        "exercises": [asdict(x) for x in seed.exercises],
        "routines": [asdict(x) for x in seed.routines],
        "routine_exercises": [asdict(x) for x in seed.routine_exercises],
        "sessions": [asdict(x) for x in seed.sessions],
        "session_sets": [asdict(x) for x in seed.session_sets],
        "bodyweight": seed.bodyweight,
    }
    OUT_PATH.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"OK -> {OUT_PATH}")
    print(f"  exercises:         {len(seed.exercises)}")
    print(f"  routines:          {len(seed.routines)}")
    print(f"  routine_exercises: {len(seed.routine_exercises)}")
    print(f"  sessions:          {len(seed.sessions)}")
    print(f"  session_sets:      {len(seed.session_sets)}")
    print(f"  bodyweight:        {len(seed.bodyweight)}")


if __name__ == "__main__":
    main()
